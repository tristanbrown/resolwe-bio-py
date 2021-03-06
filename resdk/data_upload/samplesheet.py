"""Annotation spreadsheet import, parsing, and validation."""

from __future__ import absolute_import, division, print_function, unicode_literals

import csv
import logging
import os
from collections import OrderedDict
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Protection
from openpyxl.worksheet.datavalidation import DataValidation

__all__ = ('FileImporter', 'FileExporter', )

logger = logging.getLogger(__name__)

# TODO: Construct these constants from the descriptor schema yaml files.

BASIC = [
    'SAMPLE_NAME',
    'READS_1',
    'READS_2',
    'BARCODES_FILE',
    'SEQ_TYPE',
]

SAMPLE_INFO = [
    'ANNOTATOR',
    'ORGANISM',
    'SOURCE',
    'CELL_TYPE',
    'STRAIN',
    'GENOTYPE',
    'MOLECULE',
    'DESCRIPTION',
]

PROTOCOLS = [
    'GROWTH_PROTOCOL',
    'TREATMENT_PROTOCOL',
    'EXTRACT_PROTOCOL',
    'LIBRARY_PREP',
    'FRAGMENTATION_METHOD',
]

SEQ_DATA = [
    'SEQ_DATE',
    'BARCODE_REMOVED',
]

READS_DATA = [
    'BARCODE',
    'INSTRUMENT_TYPE',
    'FACILITY',
]

ANTIBODY = ['ANTIBODY']

OPTIONAL = [
    'AGE',
    'LIBRARY_STRATEGY',
    'TISSUE',
    'OTHER_CHAR_1',
    'OTHER_CHAR_2',
]

COLUMNS = (
    BASIC
    + SAMPLE_INFO
    + PROTOCOLS
    # + SEQ_DATA # TODO: Validation incompatible with Resolwe
    + READS_DATA
    + ANTIBODY
    + OPTIONAL
)

ORGANISM = {
    'Homo sapiens',
    'Mus musculus',
    'Dictyostelium discoideum',
    'Rattus norvegicus',
}

MOLECULE = {
    'total RNA',
    'polyA RNA',
    'cytoplasmic RNA',
    'nuclear RNA',
    'genomic DNA',
    'protein',
    'other',
}

SEQ_TYPE = {
    'RNA-Seq',
    'Chemical mutagenesis',
    'miRNA-Seq',
    'ncRNA-Seq'
    'RNA-Seq (CAGE)',
    'RNA-Seq (RACE)',
    'ChIP-Seq',
    'ChIPmentation',
    'ChIP-Rx',
    'MNase-Seq',
    'MBD-Seq',
    'MRE-Seq',
    'Bisulfite-Seq',
    'Bisulfite-Seq (reduced representation)',
    'MeDIP-Seq',
    'DNase-Hypersensitivity',
    'Tn-Seq',
    'FAIRE-seq',
    'SELEX',
    'RIP-Seq',
    'ChIA-PET',
    'eClIP',
    'OTHER',
}

EMPTY = [
    '',
    'N/A',
    'NONE',
    None,
]

REQUIRED = {
    'SAMPLE_NAME',
    'SEQ_TYPE',
    'ANNOTATOR',
    'ORGANISM',
    'SOURCE',
    'MOLECULE',
}

LIMITED = {
    'SEQ_TYPE': SEQ_TYPE,
    'ORGANISM': ORGANISM,
    'MOLECULE': MOLECULE,
    'BARCODE_REMOVED': {'1', '0'},
}


class FileImporter(object):
    """Import annotation spreadsheet.

    :param str annotation_path: path to a sample annotation spreadsheet.
    """

    def __init__(self, annotation_path):
        """Validate the annotation sheet and create the sample list."""
        self._is_file(annotation_path)
        entry_list = self._populate_entries(annotation_path)
        self.valid_samples, self.invalid_names = self._create_all_samples(entry_list)

        if self.invalid_names:
            logger.error(
                "\nInvalid annotations were provided for the following samples: %s."
                "\nPlease fill in all bolded columns of the template "
                "generated by the `export_annotation` method of"
                " your collection.",
                ', '.join(self.invalid_names)
            )

    def _is_file(self, path):
        """Check is the provided path exists."""
        if not os.path.isfile(path):
            raise OSError(
                "The provided annotation file '{}' "
                "does not exist.".format(path)
            )

    def _get_spreadsheet_extension(self, path):
        """Find spreadsheet file extension."""
        return os.path.splitext(path)[1]

    def _read_xls(self, path):
        """Read Excel spreadsheet annotation file."""
        workbook = load_workbook(path)
        worksheet = workbook.active
        rows = worksheet.rows
        header = [cell.value for cell in next(rows)]
        return [self._parse_row(header, row) for row in rows]

    def _parse_row(self, header, row):
        """Convert spreadsheet row into sample entry."""
        return {head: self._parse_cell(cell) for head, cell in zip(header, row)}

    def _parse_cell(self, cell):
        """Interpret spreadsheet cell."""
        if isinstance(cell.value, float):
            return str(cell.value)
        elif cell.value in EMPTY:
            return ''
        else:
            return cell.value

    def _read_text_file(self, path):
        """Read simple spreadsheet annotation file."""
        with open(path, 'rb') as sample_sheet:
            return list(csv.DictReader(sample_sheet, delimiter='\t'))

    def _populate_entries(self, path):
        """Check the format of annotation file and assign read function."""
        if self._get_spreadsheet_extension(path) in ['.xls', '.xlsx', '.xlsm']:
            return self._read_xls(path)
        elif self._get_spreadsheet_extension(path) in ['.txt', '.tab', '.tsv']:
            return self._read_text_file(path)
        else:
            raise TypeError(
                "Annotation spreadsheet extension '{}' not recognised. Options"
                " are: '.xls', '.xlsx', '.xlsm', '.txt', '.tab', "
                "'.tsv'.".format(self._get_spreadsheet_extension(path))
            )

    def _create_all_samples(self, entries):
        """Create a sample from each samplesheet entry."""
        valid_samples = OrderedDict()
        invalid_names = set()
        for entry in entries:
            self._create_sample(entry, valid_samples, invalid_names)
        return valid_samples, invalid_names

    def _create_sample(self, entry, valid_samples, invalid_names):
        """Create a sample from a samplesheet entry, if unique."""
        name = entry['SAMPLE_NAME']
        if name in invalid_names or name in valid_samples:
            logger.error(
                "The sample name '%s' is duplicated. Please use unique "
                "sample names.",
                name
            )
            invalid_names.add(name)
            valid_samples.pop(name, None)
        else:
            try:
                valid_samples[name] = Sample(entry)
            except ValueError as ex:
                invalid_names.add(name)
                logger.error(ex)


class FileExporter(object):
    """Export annotation spreadsheet.

    :param str annotation_path: path to write the sample annotation spreadsheet
    :param sample_list: a list of resdk sample objects
    """

    def __init__(self, sample_list=[], export_path=None):
        """Initialize the samplesheet template."""
        self.path = export_path
        self._samples = sample_list
        self._template = self._create_template(COLUMNS)
        for sample in sample_list:
            self._add_entry(sample)
        self._template.save(filename=self.path)

    def _create_template(self, headers):
        """Construct a template samplesheet."""
        template = Workbook()
        sheet = template.active

        # Add headers and lock the sheet
        sheet.append(headers)
        sheet.protection.sheet = True

        # Apply formats to everything
        for cell in sheet[1]:
            self._apply_xlsm_formats(sheet, cell)

        # Return the template
        return template

    def _apply_xlsm_formats(self, sheet, cell):
        """Apply column-specific, macro-enabled spreadsheet formats."""
        # Create styles
        normal = Font(name='Arial')
        bold = copy(normal)
        bold.bold = True

        # Acquire indices and headers
        header = cell.value
        col_id = cell.column
        col = sheet.column_dimensions[col_id]
        col.font = normal
        col.width = self._get_column_width(header)

        # Lock only the headers
        col.protection = Protection(locked=False)
        cell.font = normal  # Required for locking (openpyxl bug?)

        # Format the required columns
        if header in REQUIRED:
            cell.font = bold

        # Format the columns with limited options
        try:
            options = '"{}"'.format(','.join(LIMITED[header]))
            valid = DataValidation(type="list", formula1=options)
            valid.error = "Invalid {}.".format(header)
            sheet.add_data_validation(valid)
            valid.add(self._get_column_body(col_id))
            col.width = self._get_column_width(LIMITED[header])
        except KeyError:
            pass

        # Format the date column
        if header == 'SEQ_DATE':
            valid_date = DataValidation(type="date")
            valid_date.error = "Invalid date."
            sheet.add_data_validation(valid_date)
            valid_date.add(self._get_column_body(col_id))

    def _get_column_body(self, column):
        """Give the indices for the entire column, minus the header."""
        return '{0}2:{0}1048576'.format(column)

    def _get_column_width(self, words, factor=1.7, limits=(8, 20)):
        """Choose a column width based on the given list of words."""
        if isinstance(words, str):
            words = [words]

        width = factor * max([len(word) for word in words])

        if width > limits[1]:
            width = limits[1]
        elif width < limits[0]:
            width = limits[0]

        return width

    def _add_entry(self, sample):
        """Add a sample as an entry to the samplesheet."""
        sheet = self._template.active

        annotation = self._extract_descriptors(sample)
        entry = [annotation.get(header, '') for header in COLUMNS]

        sheet.append(entry)

    def _extract_descriptors(self, sample):
        """Extract all sample annotation info as a dictionary."""
        # Populate the sample info
        info = {'SAMPLE_NAME': sample.name}
        if sample.descriptor:
            info.update(sample.descriptor['sample'])
            info.update(self._extract_optional(info.pop('optional_char', [])))

        # Populate the raw sequencing characteristics
        try:
            reads = sample.data.filter(type='data:reads')
            info.update(self._extract_seqinfo(reads[0].descriptor))
        except IndexError:
            logger.warning("No reads found for sample '%s'.", sample.name)
        except KeyError:
            logger.warning("Sample '%s' reads not annotated.", sample.name)

        # Eliminate null values and capitalize headers
        return {key.upper(): '' if val in EMPTY else val for key, val in info.items()}

    def _extract_optional(self, char_list):
        """Convert a list of optional characteristics into a dictionary."""
        return dict(char.split(':') for char in char_list)

    def _extract_seqinfo(self, info):
        """Extract reads annotation info from a sample."""
        entry = {'SEQ_TYPE': info['experiment_type']}
        if 'reads_info' in info:
            entry.update(_dict_upper(info['reads_info']))
        if 'protocols' in info:
            entry.update(_dict_upper(info['protocols']))
        return entry


class Sample(object):
    """Create a Sample like object.

    :param dict entry: a dictionary containing header:data pairs generated from
        an annotation spreadsheet
    """

    # TODO: Abstract this to handle other descriptor schema types.
    def __init__(self, entry):
        """Validate the entry and construct the sample descriptor."""
        self.valid = self.validate(entry)
        self._build_descriptors(entry)
        self.community_tag = self._get_community_tag(self.seq_type.lower())

    def _build_descriptors(self, entry):
        """Extract the sample meta-data."""
        self.name = entry['SAMPLE_NAME']
        self.path = entry['READS_1']
        self.path2 = entry['READS_2']
        self.path3 = entry['BARCODES_FILE']
        self.seq_type = entry['SEQ_TYPE']
        self.barcode = entry['BARCODE']

        # Build reads descriptor
        protocols = {char.lower(): entry[char] for char in PROTOCOLS}
        antibody = {
            'antibody_information': {'manufacturer': entry['ANTIBODY']}
        }
        protocols.update(antibody)
        reads_info = {char.lower(): entry[char] for char in READS_DATA}

        self.reads_annotation = {'protocols': protocols, 'reads_info': reads_info}

        # TODO: Fix format incompatibility between openpyxl and Resolwe
        # for char in SEQ_DATA:
        #     if entry[char]:
        #         self.reads_annotation['reads_info'][char.lower()] = entry[char]

        # Build remaining sample descriptor
        self.molecule = entry['MOLECULE']
        self.organism = entry['ORGANISM']
        self.annotator = entry['ANNOTATOR']
        self.source = entry['SOURCE']
        self.sample_annotation = {
            'sample': {
                'cell_type': entry['CELL_TYPE'],
                'strain': entry['STRAIN'],
                'genotype': entry['GENOTYPE'],
                'description': entry['DESCRIPTION'],
            }
        }

        # Include only if they are non-empty, to not override error-checking
        if self.seq_type:
            self.reads_annotation['experiment_type'] = self.seq_type

        fields = [
            ('organism', self.organism),
            ('molecule', self.molecule),
            ('annotator', self.annotator),
            ('source', self.source),
        ]
        reqfields = {label: info for label, info in fields if info}
        self.sample_annotation['sample'].update(reqfields)

        # Include optional columns
        optional = [
            '{0}:{1}'.format(char, entry[char])
            for char in sorted(OPTIONAL) if entry[char]
        ]
        self.sample_annotation['sample']['optional_char'] = optional

    def validate(self, entry):
        """Validate the annotation spreadsheet file."""
        # Check column headers
        diff1 = set(COLUMNS) - set(entry.keys())
        diff2 = set(entry.keys()) - set(COLUMNS)
        err_head = (
            "Headers '{}' {}. You should use the headers generated by"
            " the `export_annotation` method of your collection."
        )
        if diff1:
            raise KeyError(
                err_head.format("', '".join(diff1), "are missing")
            )
        if diff2:
            raise KeyError(
                err_head.format("', '".join(diff2), "not recognized")
            )

        # Check required, restricted values
        err_req = "For the sample, '{}', '{}' is not a valid {}."

        restricted = [
            ('organism', ORGANISM),
            ('molecule', MOLECULE),
            ('seq_type', SEQ_TYPE),
        ]
        for var_name, options in restricted:
            var = entry[var_name.upper()]
            if var not in options:
                raise ValueError(
                    err_req.format(
                        entry['SAMPLE_NAME'], var, var_name.upper()
                    )
                )

        # Check required, unrestricted values
        for var_name in ['sample_name', 'annotator', 'source']:
            var = entry[var_name.upper()]
            if var.upper() in EMPTY:
                raise ValueError(
                    err_req.format(
                        entry['SAMPLE_NAME'], var, var_name.upper()
                    )
                )

    def _get_community_tag(self, experiment):
        """Prepare community tags."""
        if 'rna' in experiment:
            community = 'community:rna-seq'
        elif 'chip' in experiment:
            community = 'community:chip-seq'
        elif 'chemical' in experiment:
            community = 'community:dicty'
        else:
            community = None
        return community


def _dict_upper(a_dict):
    """Capitalizes the keys of a dictionary."""
    return {key.upper(): value for key, value in a_dict.items()}
